# -*- coding: utf-8 -*-
"""
build_5tools_delivery.py
余嘉 5 工具交付包 —— 对齐袁老师样例 data/Sample_merged_prime_results.xlsx 的形式。
- backbone = 样例前 15 列（DS2 ELISpot，33922 行），每张表行对行与 PRIME 表对齐
- 每工具一张 xlsx：backbone + 该工具原生输出列（不同工具列数不同）
- 第二 sheet「列说明」解释新增列含义
join key = (MT_Subpeptide, HLA_Allele) MT 侧 / (WT_Subpeptide, HLA_Allele) WT 侧；NeoTImmuML 肽无关按 subpeptide。
覆盖不到的行留空（工具肽长/HLA 限制，诚实 NA）。
"""
import openpyxl, csv, re, os
from openpyxl import Workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SAMPLE = os.path.join(ROOT, 'data', 'Sample_merged_prime_results.xlsx')
OUT = os.path.join(ROOT, '5tools_delivery')
os.makedirs(OUT, exist_ok=True)

BB_COLS = ['Patient_ID','Peptide_ID','Treatment','Gene_Name','Mutation','Vaccine_Peptide',
           'WT_Peptide_Seq','Peptide_Length','TPM_PurifiedTumorRNA','Elispot','Window_Size',
           'Position','MT_Subpeptide','WT_Subpeptide','HLA_Allele']

def load_backbone():
    wb = openpyxl.load_workbook(SAMPLE, read_only=True); ws = wb.active
    hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    ix = {h:i for i,h in enumerate(hdr)}
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        rows.append({c: r[ix[c]] for c in BB_COLS})
    wb.close()
    return rows

def norm_hla_star(hla):
    # HLA-A24:02 -> HLA-A*24:02 ; already-canonical passthrough
    if hla is None: return None
    m = re.match(r'^(HLA-[A-Z])(\d+:\d+)$', hla)
    return '%s*%s' % (m.group(1), m.group(2)) if m else hla

# ---------- per-tool native column maps ----------
def map_from_merged(path, mt_cols, wt_cols):
    """从已 HLA-映射好的 merged_<tool>.xlsx 取原生列。key MT 侧 (MT_Subpeptide,HLA)，WT 侧 (WT_Subpeptide,HLA)。"""
    wb = openpyxl.load_workbook(path, read_only=True); ws = wb.active
    hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    ix = {h:i for i,h in enumerate(hdr)}
    mt, wt = {}, {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        hla = r[ix['HLA_Allele']]
        kmt = (r[ix['MT_Subpeptide']], hla)
        kwt = (r[ix['WT_Subpeptide']], hla)
        if kmt not in mt:
            mt[kmt] = {c: r[ix[c]] for c in mt_cols}
        if kwt not in wt:
            wt[kwt] = {c: r[ix[c]] for c in wt_cols}
    wb.close()
    return mt, wt

def build(tool, out_cols, fill_fn, dictionary):
    bb = load_backbone()
    wb = Workbook(); ws = wb.active; ws.title = tool
    ws.append(BB_COLS + out_cols)
    n_hit = 0
    for row in bb:
        extra = fill_fn(row)
        if any(extra.get(c) not in (None, '') for c in out_cols):
            n_hit += 1
        ws.append([row[c] for c in BB_COLS] + [extra.get(c) for c in out_cols])
    # 列说明 sheet
    ds = wb.create_sheet('列说明')
    ds.append(['列名', '含义'])
    for k, v in dictionary.items():
        ds.append([k, v])
    path = os.path.join(OUT, '%s.xlsx' % tool)
    wb.save(path)
    print('[%s] %d rows, %d/%d backbone rows scored -> %s' % (tool, len(bb), n_hit, len(bb), path))

# ===== 1. PredIG (merged_predig，原生特征全) =====
PREDIG_MT = ['MT_PredIG','MT_NOAH','MT_NetCleave','MT_Stab_peptide','MT_TCR_contact',
             'MT_Hydrophobicity_peptide','MT_MW_peptide','MT_Charge_peptide',
             'MT_Hydrophobicity_tcr_contact','MT_MW_tcr_contact','MT_Charge_tcr_contact']
PREDIG_WT = ['WT_PredIG','WT_NOAH','WT_NetCleave','WT_Stab_peptide','WT_TCR_contact',
             'WT_Hydrophobicity_peptide','WT_MW_peptide','WT_Charge_peptide',
             'WT_Hydrophobicity_tcr_contact','WT_MW_tcr_contact','WT_Charge_tcr_contact']
pmt, pwt = map_from_merged(os.path.join(ROOT,'scripts','out','merged_predig.xlsx'), PREDIG_MT, PREDIG_WT)
def fill_predig(row):
    d = {}
    d.update(pmt.get((row['MT_Subpeptide'], row['HLA_Allele']), {}))
    d.update(pwt.get((row['WT_Subpeptide'], row['HLA_Allele']), {}))
    return d
build('PredIG', PREDIG_MT + PREDIG_WT, fill_predig, {
    'MT_PredIG':'PredIG-Neo 突变肽免疫原性总分（越高越可能免疫原，主输出）',
    'WT_PredIG':'对应野生肽免疫原性总分',
    'MT_NOAH/WT_NOAH':'NOAH pMHC 结合预测分量',
    'MT_NetCleave/WT_NetCleave':'NetCleave 蛋白酶切割位点分量',
    'MT_Stab_peptide/WT_Stab_peptide':'pMHC 稳定性分量',
    'MT_TCR_contact/WT_TCR_contact':'TCR 接触面特征分量',
    'MT_Hydrophobicity_peptide/...':'全肽疏水性',
    'MT_MW_peptide/...':'全肽分子量',
    'MT_Charge_peptide/...':'全肽电荷',
    'MT_*_tcr_contact':'TCR 接触残基的疏水性/分子量/电荷',
    '空白':'该肽×HLA 不在 PredIG 支持范围（肽长/HLA），无分',
})

# ===== 2. DeepImmuno (merged_deepimmuno，单免疫原性分) =====
dmt, dwt = map_from_merged(os.path.join(ROOT,'scripts','out','merged_deepimmuno.xlsx'),
                           ['MT_DeepImmuno'], ['WT_DeepImmuno'])
def fill_deep(row):
    d = {}
    d.update(dmt.get((row['MT_Subpeptide'], row['HLA_Allele']), {}))
    d.update(dwt.get((row['WT_Subpeptide'], row['HLA_Allele']), {}))
    return d
build('DeepImmuno', ['MT_DeepImmuno','WT_DeepImmuno'], fill_deep, {
    'MT_DeepImmuno':'突变肽免疫原性概率 0-1（CNN，越高越免疫原，主输出）',
    'WT_DeepImmuno':'野生肽免疫原性概率 0-1',
    '空白':'仅支持 9-10mer，其它肽长无分',
})

# ===== 3. pTuneos (raw ptuneos_unique_output.tsv，HLA 已规范) =====
PT_COLS = ['MT_pTuneos','pTuneos_Recognition_score','pTuneos_Hydrophobicity_score',
           'pTuneos_Self_sequence_similarity','pTuneos_MT_Binding_EL','pTuneos_WT_Binding_EL',
           'pTuneos_hydro_defaulted']
pt = {}
with open(os.path.join(ROOT,'scripts','ptuneos','ptuneos_unique_output.tsv')) as f:
    for r in csv.DictReader(f, delimiter='\t'):
        pt[(r['MT_pep'], r['HLA_type'])] = {
            'MT_pTuneos': r['model_pro'],
            'pTuneos_Recognition_score': r['Recognition_score'],
            'pTuneos_Hydrophobicity_score': r['Hydrophobicity_score'],
            'pTuneos_Self_sequence_similarity': r['Self_sequence_similarity'],
            'pTuneos_MT_Binding_EL': r['MT_Binding_EL'],
            'pTuneos_WT_Binding_EL': r['WT_Binding_EL'],
            'pTuneos_hydro_defaulted': r['hydro_defaulted'],
        }
def fill_pt(row):
    return pt.get((row['MT_Subpeptide'], row['HLA_Allele']), {})
build('pTuneos', PT_COLS, fill_pt, {
    'MT_pTuneos':'Pre&RecNeo 免疫原性总概率 model_pro 0-1（主输出，越高越免疫原）',
    'pTuneos_Recognition_score':'T 细胞识别打分',
    'pTuneos_Hydrophobicity_score':'TCR 接触位疏水性打分',
    'pTuneos_Self_sequence_similarity':'与自身蛋白组相似度（越低越非己）',
    'pTuneos_MT_Binding_EL':'突变肽 netMHCpan %Rank EL（越小越强结合）',
    'pTuneos_WT_Binding_EL':'野生肽 netMHCpan %Rank EL',
    'pTuneos_hydro_defaulted':'疏水分是否用了默认填补（True=该肽缺值兜底）',
    '空白':'该肽×HLA 未跑出（肽长/HLA 限制）',
})

# ===== 4. IMPROVE (raw improve_full_result.tsv，HLA 加星号；MT-only RF) =====
IMP_RAW = ['mean_prediction_rf','RankEL','RankBA','RankEL_wt','Stability','Prime','DAI',
           'SelfSim','Expression','Foreigness','HydroAll','HydroCore']
IMP_COLS = ['MT_IMPROVE_'+c for c in IMP_RAW]
imp = {}
with open(os.path.join(ROOT,'scripts','out','improve_full_result.tsv')) as f:
    for r in csv.DictReader(f, delimiter='\t'):
        key = (r['Mut_peptide'], norm_hla_star(r['HLA_allele']))
        imp[key] = {'MT_IMPROVE_'+c: r.get(c) for c in IMP_RAW}
def fill_imp(row):
    return imp.get((row['MT_Subpeptide'], row['HLA_Allele']), {})
build('IMPROVE', IMP_COLS, fill_imp, {
    'MT_IMPROVE_mean_prediction_rf':'IMPROVE 随机森林集成最终免疫原性分（主输出，越高越免疫原）',
    'MT_IMPROVE_RankEL/RankBA':'netMHCpan EL/BA %Rank（结合强度，越小越强）',
    'MT_IMPROVE_RankEL_wt':'野生肽 EL %Rank',
    'MT_IMPROVE_Stability':'netMHCstabpan 稳定性',
    'MT_IMPROVE_Prime':'PRIME 免疫原性分量',
    'MT_IMPROVE_DAI':'差异凝集指数（MT vs WT 结合差）',
    'MT_IMPROVE_SelfSim':'自身相似度',
    'MT_IMPROVE_Expression':'基因表达（TPM 派生）',
    'MT_IMPROVE_Foreigness':'外源性（与已知抗原相似度）',
    'MT_IMPROVE_HydroAll/HydroCore':'全肽/核心区疏水性',
    '说明':'IMPROVE 是突变肽 RF 集成，只对 MT 打分（无 WT 列）；Expression 特征本部署为降级值',
    '空白':'该肽×HLA 不在 netMHC 支持肽长/HLA 内，无分',
})

# ===== 5. NeoTImmuML (raw neotimmuml_scores.csv，肽无关 HLA-agnostic) =====
neo = {}
with open(os.path.join(ROOT,'scripts','out','neotimmuml_scores.csv')) as f:
    for r in csv.DictReader(f):
        neo[r['Peptide']] = r
NEO_COLS = ['MT_NeoTImmuML','WT_NeoTImmuML','MT_NeoTImmuML_rf_proba',
            'MT_NeoTImmuML_lgb_proba','MT_NeoTImmuML_xgb_proba']
def fill_neo(row):
    d = {}
    m = neo.get(row['MT_Subpeptide']); w = neo.get(row['WT_Subpeptide'])
    if m:
        d['MT_NeoTImmuML'] = m['neotimmuml_score']
        d['MT_NeoTImmuML_rf_proba'] = m['rf_proba']
        d['MT_NeoTImmuML_lgb_proba'] = m['lgb_proba']
        d['MT_NeoTImmuML_xgb_proba'] = m['xgb_proba']
    if w:
        d['WT_NeoTImmuML'] = w['neotimmuml_score']
    return d
build('NeoTImmuML', NEO_COLS, fill_neo, {
    'MT_NeoTImmuML':'突变肽免疫原性总分（RF+LGB+XGB 集成，主输出，越高越免疫原）',
    'WT_NeoTImmuML':'野生肽免疫原性总分',
    'MT_NeoTImmuML_rf_proba/lgb_proba/xgb_proba':'三个基模型各自概率（集成成分）',
    '说明':'NeoTImmuML 仅用肽理化特征，HLA-agnostic（不分等位基因，同肽各 HLA 同分）；自训复刻版（官方权重不可得，PPT 标★）',
    '空白':'该肽长不在模型支持范围',
})

print('DONE ->', OUT)
