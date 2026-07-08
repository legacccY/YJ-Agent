#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
reconstruct_wt_official.py
服务: quantimmu-bench / Phase 0  (03_EXPERIMENT_PLAN.md §3 + 消融 AB-6)

补 DAI(MT vs WT) 的 WT 侧:
  新官方数据只有 MT 候选肽(Vaccine_Peptide), 无显式 WT 列。
  本脚本从突变记法回推 WT 全长肽(仅 SNV), 验证方法对不对(对旧 xlsx 金标准),
  再把 WT 全长滑窗成 WT 子肽×HLA (与 MT 同口径) 供工具补跑 WT 侧分数。

  indel(DEL/INS): WT 本无定义 -> 标 N/A, 不产 WT 子肽。
  (Variant_Type 缺失但 p.XnY 记法可解析的行, 如 AMACR 16097-104-24, 视为 SNV; 见 is_snv。)

================== 输入 (只读 frozen + 官方) ==================
  data/frozen/ds2_official_groundtruth.csv   (130 肽; 列 Vaccine_Peptide=MT 全长,
       Short_Epitope, Gene_and_Protein_Change=GENE|p.XnY, Variant_Type, ...)
  data/frozen/RERUN_PEPTIDE_LIST.csv         (43 待补跑肽 + hla_alleles_to_run)
  data/frozen/REUSE_DECISION.csv             (status reuse/rerun_full/rerun_partial)
  data/frozen/patient_hla.csv                (std 等位真源; hla_alleles_to_run 即由此来)
  scripts/out/merged_all_tools_29tools.xlsx  (Sheet1; 旧预测金标准, 含
       MT_FullPeptide + WT_FullPeptide; 用 87 reuse 肽里 SNV 交叉验证回推方法)

================== WT 来源优先级 (仅 SNV) ==================
  1. 优先 (gold_reuse): 肽(按 Peptide_ID)在旧 xlsx 有 WT_FullPeptide 金标准
     -> 直接取旧金标准 (已验证 100% 可信, 含 41 个回推歧义肽如 P104)。
  2. 回退 (derived): 真·新肽(不在旧 xlsx, rerun_full) 且 SNV -> 用 p.XnY 回推:
       Gene_and_Protein_Change = GENE|p.{WT_aa}{prot_pos}{MT_aa}  e.g. PIK3CA|p.E545K
         -> wt_aa=E, mt_aa=K  (prot_pos=545 是蛋白坐标, 不可直接索引长肽!)
       在 Vaccine_Peptide(MT 全长) 内定位突变位 (1-based):
         主法: mt_aa 在长肽中唯一出现 -> 该位
         歧义: mt_aa 多次出现 -> 用 Short_Epitope(含突变残基)在长肽中的 span 缩小,
                只保留落在表位 span 内的 mt_aa 位; 唯一即定位, 否则标 ambiguous(不臆造)
       WT 全长 = MT 把该位 mt_aa 换回 wt_aa。
  3. indel(DEL/INS) 或记法非干净替换 -> indel_NA (WT 无定义, 不变)。

  注: derived 回推法仍对旧金标准独立交叉验证 (见 [WT-g2], 凡有 gold 的 SNV
      都跑回推并比对, 证明回推法正确; gold_reuse 输出天然=旧值无需比)。

================== 输出 ==================
  data/frozen/wt_fullpeptide_official.csv
    列: mut_key, Peptide_ID, Vaccine_Peptide_MT, WT_FullPeptide,
        mut_pos, wt_aa, mt_aa, status[gold_reuse/derived/ambiguous/indel_NA]
  data/frozen/subpep_hla_expansion_WT.csv
    列对齐 MT 版 + WT_FullPeptide + side='WT':
        mut_key, Patient_ID, Peptide_ID, Vaccine_Peptide(=MT 全长, 溯源),
        WT_FullPeptide, subpep_seq(WT 子肽), subpep_pos, window_size,
        hla_allele_std, side
  data/frozen/WT_NA_indel_list.csv
    所有无 WT 的肽(DEL/INS + Variant_Type 缺): mut_key, Peptide_ID, Variant_Type,
        Gene_and_Protein_Change, status

================== 校验门 ==================
  [WT-g1] SNV 数 == 102  且 gold_reuse_SNV + derived + ambiguous == SNV 数
  [WT-g2] derived 回推法 vs 旧 xlsx 金标准 match 率打印; <100% 打印失配清单(不 silent)
  [WT-g3] indel/无 WT 数 == 28
  [WT-g4] gold_reuse / derived / ambiguous(理想 0) / indel_NA 各数打印

================== 跑法 (不在本脚本内跑; 交主线) ==================
  python scripts/reconstruct_wt_official.py            # 9mer 默认 (与 MT 同)
  python scripts/reconstruct_wt_official.py --window 8-11
"""

import re
import sys
import argparse
from pathlib import Path
from collections import defaultdict

import pandas as pd
import openpyxl

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
FROZEN_DIR = ROOT / "data" / "frozen"

GT_CSV = FROZEN_DIR / "ds2_official_groundtruth.csv"
RERUN_CSV = FROZEN_DIR / "RERUN_PEPTIDE_LIST.csv"
REUSE_CSV = FROZEN_DIR / "REUSE_DECISION.csv"
OLD_XLSX = ROOT / "scripts" / "out" / "merged_all_tools_29tools.xlsx"

OUT_WT_FULL = FROZEN_DIR / "wt_fullpeptide_official.csv"
OUT_WT_SUBPEP = FROZEN_DIR / "subpep_hla_expansion_WT.csv"
OUT_NA_LIST = FROZEN_DIR / "WT_NA_indel_list.csv"

# 复用 MT 同口径滑窗逻辑 (别重写, 保证 MT/WT 同 9mer 窗)
sys.path.insert(0, str(ROOT / "analysis" / "phase0"))
from p0c_subpep_expansion import slide, parse_window  # noqa: E402

# 干净 SNV 替换记法: p.{WT}{pos}{MT}, 单字母氨基酸
SNV_RE = re.compile(r"^p\.([A-Z])(\d+)([A-Z])$")


# ─────────────────────────────────────────────────────────────────────────
# 解析突变记法
# ─────────────────────────────────────────────────────────────────────────
def parse_protein_change(gene_change):
    """'PIK3CA|p.E545K' -> ('E','K',545); 非干净 SNV 替换 -> None。"""
    if gene_change is None or (isinstance(gene_change, float)):
        return None
    s = str(gene_change).strip()
    if "|" in s:
        s = s.split("|", 1)[1]
    s = s.strip()
    m = SNV_RE.match(s)
    if not m:
        return None
    wt_aa, prot_pos, mt_aa = m.group(1), int(m.group(2)), m.group(3)
    return wt_aa, mt_aa, prot_pos


# ─────────────────────────────────────────────────────────────────────────
# 在 MT 全长肽内定位突变位 (1-based), 回推 WT
# ─────────────────────────────────────────────────────────────────────────
def locate_mut_pos(mt_pep, mt_aa, short_epitope):
    """
    返回 (pos_1based, status):
      status='ok' 唯一定位; 'ambiguous' 多解/找不到 (不臆造)。
    主法: mt_aa 在 mt_pep 唯一出现。
    歧义: 用 short_epitope 在 mt_pep 的 span 缩小, 只留落在 span 内的 mt_aa 位。
    """
    mt_pep = "" if mt_pep is None else str(mt_pep).strip()
    occ = [i for i, ch in enumerate(mt_pep) if ch == mt_aa]  # 0-based
    if len(occ) == 0:
        return None, "ambiguous"  # mt_aa 不在长肽 -> 记法/肽不符, 不臆造
    if len(occ) == 1:
        return occ[0] + 1, "ok"

    # 多次出现 -> 用表位 span 缩小
    epi = "" if short_epitope is None or isinstance(short_epitope, float) else str(short_epitope).strip()
    if epi:
        starts = [m.start() for m in re.finditer(re.escape(epi), mt_pep)]
        # 表位在长肽唯一定位时才用其 span 过滤 (表位本身多解则放弃缩小)
        if len(starts) == 1:
            lo = starts[0]
            hi = lo + len(epi)  # [lo, hi)
            in_span = [p for p in occ if lo <= p < hi]
            if len(in_span) == 1:
                return in_span[0] + 1, "ok"
    return None, "ambiguous"


def reconstruct_wt(mt_pep, wt_aa, mt_aa, pos_1based):
    """MT 全长把 pos 处 mt_aa 换回 wt_aa -> WT 全长。pos 处必须确为 mt_aa。"""
    mt_pep = str(mt_pep).strip()
    i = pos_1based - 1
    if not (0 <= i < len(mt_pep)) or mt_pep[i] != mt_aa:
        return None
    return mt_pep[:i] + wt_aa + mt_pep[i + 1:]


# ─────────────────────────────────────────────────────────────────────────
# 旧 xlsx 金标准: 官方患者(Patient_ID>=100) 的 Peptide_ID -> WT_FullPeptide
# ─────────────────────────────────────────────────────────────────────────
def load_old_wt_map(xlsx_path):
    """读 Sheet1, 取 Patient_ID>=100 的 Peptide_ID -> WT_FullPeptide (peptide 级唯一)。"""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = ws.iter_rows(min_row=1, values_only=True)
    hdr = next(rows)
    idx = {h: i for i, h in enumerate(hdr)}
    wt_by_pep = defaultdict(set)
    for r in rows:
        pat = r[idx["Patient_ID"]]
        if pat is None or pat < 100:  # 只取新官方患者
            continue
        pep = str(r[idx["Peptide_ID"]])
        wt = r[idx["WT_FullPeptide"]]
        if wt is not None and str(wt).strip():
            wt_by_pep[pep].add(str(wt).strip())
    wb.close()
    # peptide 级取唯一值 (已核 0 个多解)
    out = {}
    for pep, vals in wt_by_pep.items():
        out[pep] = sorted(vals)[0] if len(vals) == 1 else None  # None=多解, 验证时跳并报
    multi = [p for p, v in wt_by_pep.items() if len(v) > 1]
    return out, multi


def main():
    ap = argparse.ArgumentParser(description="官方数据 WT 全长回推 + WT 子肽×HLA 展开 (不跑工具)")
    ap.add_argument("--window", default="9",
                    help="滑窗口径: '9' (默认, 与 MT 同) 或 '8-11' (补充)")
    args = ap.parse_args()
    windows = parse_window(args.window)
    print(f"[info] 滑窗口径: {windows}")

    for p in (GT_CSV, RERUN_CSV, REUSE_CSV):
        if not p.exists():
            raise SystemExit(f"[ERR] 依赖缺失: {p}")

    gt = pd.read_csv(GT_CSV)
    print(f"[info] 官方肽数: {len(gt)}")

    # ── 0. 预载旧 xlsx 金标准 WT (Peptide_ID -> WT_FullPeptide; gold-first 来源) ─
    old_wt, old_multi = {}, []
    if OLD_XLSX.exists():
        old_wt, old_multi = load_old_wt_map(OLD_XLSX)
        print(f"[info] 旧 xlsx 金标准 WT 肽数: {sum(1 for v in old_wt.values() if v)}")
        if old_multi:
            print(f"[WARN] 旧 xlsx 有 {len(old_multi)} 肽 WT 多解(将不作 gold, 回退回推): {sorted(old_multi)}")
    else:
        print(f"[WARN] 旧金标准缺失, 全部走回推(无 gold_reuse): {OLD_XLSX}")

    # ── 1. 定 WT 来源 (逐肽; gold 优先, 真新肽才回推) ─────────────────────
    full_rows = []
    na_rows = []
    # n_gold_reuse=旧金标准复用; n_derived=真新肽回推成功; n_amb=真新肽回推歧义
    n_snv = n_gold = n_derived = n_amb = 0
    derived_for_check = {}  # Peptide_ID -> 独立回推值(凡 SNV 能回推就存, 供 [WT-g2] 验证)
    for r in gt.itertuples(index=False):
        mut_key = r.mut_key
        pep_id = str(r.Peptide_ID)
        mt_pep = "" if pd.isna(r.Vaccine_Peptide) else str(r.Vaccine_Peptide).strip()
        vtype = "" if pd.isna(r.Variant_Type) else str(r.Variant_Type).strip()
        gene_change = r.Gene_and_Protein_Change
        parsed = parse_protein_change(gene_change)

        # AMACR fix: Variant_Type 可能为空但记法可解析 (如 AMACR|p.Y41N, 16097-104-24),
        # 别只信 Variant_Type 列把它误挡进 indel; 记法能被 p.XnY 解析即视为 SNV。
        is_snv = (vtype == "SNV") or (parsed is not None)
        if not is_snv or parsed is None:
            # 非 SNV (DEL/INS/Variant_Type 缺) 或记法非干净替换 -> WT 无定义
            na_rows.append({
                "mut_key": mut_key,
                "Peptide_ID": pep_id,
                "Variant_Type": vtype if vtype else "NA",
                "Gene_and_Protein_Change": "" if pd.isna(gene_change) else str(gene_change),
                "status": "indel_NA",
            })
            full_rows.append({
                "mut_key": mut_key, "Peptide_ID": pep_id,
                "Vaccine_Peptide_MT": mt_pep, "WT_FullPeptide": "",
                "mut_pos": "", "wt_aa": "", "mt_aa": "", "status": "indel_NA",
            })
            continue

        n_snv += 1
        wt_aa, mt_aa, _prot_pos = parsed

        # (a) 独立回推一次 (无论 gold 与否; gold_reuse 肽也回推, 留作 [WT-g2] 方法验证)
        pos, loc_status = locate_mut_pos(mt_pep, mt_aa, r.Short_Epitope)
        derived_wt = None
        derived_pos = ""
        if loc_status == "ok" and pos is not None:
            cand = reconstruct_wt(mt_pep, wt_aa, mt_aa, pos)
            if cand is not None:
                derived_wt = cand
                derived_pos = pos
                derived_for_check[pep_id] = cand

        # (b) 定来源: 旧金标准优先 -> 回退回推 -> 歧义
        gold = old_wt.get(pep_id)
        if gold:
            wt_full = gold
            status = "gold_reuse"
            mut_pos_out = derived_pos  # 回推到位则一并记, 否则空(gold 不依赖定位)
            n_gold += 1
        elif derived_wt is not None:
            wt_full = derived_wt
            status = "derived"
            mut_pos_out = derived_pos
            n_derived += 1
        else:
            wt_full = ""
            status = "ambiguous"
            mut_pos_out = ""
            n_amb += 1

        full_rows.append({
            "mut_key": mut_key, "Peptide_ID": pep_id,
            "Vaccine_Peptide_MT": mt_pep, "WT_FullPeptide": wt_full,
            "mut_pos": mut_pos_out, "wt_aa": wt_aa, "mt_aa": mt_aa, "status": status,
        })

    full_df = pd.DataFrame(full_rows)
    na_df = pd.DataFrame(na_rows)

    print(f"\n[WT-g1] SNV 数(实际): {n_snv}  (期望 102)  "
          f"gold_reuse={n_gold} derived={n_derived} ambiguous={n_amb}  "
          f"(和={n_gold + n_derived + n_amb})")
    print(f"[WT-g3] 无 WT(indel) 数(实际): {len(na_df)}  (期望 28)")
    print(f"[WT-g4] gold_reuse={n_gold}  derived={n_derived}  "
          f"ambiguous={n_amb}(理想 0)  indel_NA={len(na_df)}")
    if n_amb:
        amb = full_df[full_df["status"] == "ambiguous"][["mut_key", "Peptide_ID", "wt_aa", "mt_aa"]]
        print("[WT-g4] ambiguous 清单 (真新肽里残基多次出现, 未臆造定位):")
        print(amb.to_string(index=False))

    # ── 2. 交叉验证: 凡有 gold 的 SNV, 独立回推值 vs 旧 xlsx 金标准 ────────
    #      (验回推法正确性; 对所有 derived_for_check∩有 gold 的肽比, 应 100%。
    #       gold_reuse 输出天然=gold 无需比, 此处只是 sanity-check 回推法本身)
    print("\n[WT-g2] === 回推法交叉验证 (回推值 vs 旧 xlsx 金标准) ===")
    if not OLD_XLSX.exists():
        print(f"[WARN] 旧金标准缺失, 跳过验证: {OLD_XLSX}")
    else:
        checked, match, mism = 0, 0, []
        for pep_id in sorted(derived_for_check):
            gold = old_wt.get(pep_id)
            if gold is None:
                continue  # 无 gold 的真新肽无法验
            checked += 1
            mine = derived_for_check[pep_id]
            if mine == gold:
                match += 1
            else:
                mism.append((pep_id, mine, gold))
        rate = (match / checked * 100.0) if checked else float("nan")
        print(f"[WT-g2] 验证肽数(SNV∩可回推∩有金标准): {checked}  match={match}  "
              f"match 率={rate:.2f}%")
        if mism:
            print(f"[WT-g2] !!! {len(mism)} 条失配 (match 率<100%, 人工核, 不 silent 通过):")
            for pep_id, mine, gold in mism:
                print(f"    {pep_id}\n      mine={mine}\n      gold={gold}")
        else:
            print("[WT-g2] 回推法全部匹配 (100%)" if checked else "[WT-g2] 无可验肽")

    # ── 3. WT 子肽×HLA 展开 (43 待补跑肽里有 WT 的; gold_reuse+derived 都产) ──
    rerun = pd.read_csv(RERUN_CSV)
    rerun["Patient_ID"] = rerun["Patient_ID"].astype(int)
    rerun["Peptide_ID"] = rerun["Peptide_ID"].astype(str)
    wt_full_map = dict(zip(full_df["Peptide_ID"].astype(str), full_df["WT_FullPeptide"]))
    status_map = dict(zip(full_df["Peptide_ID"].astype(str), full_df["status"]))

    HAS_WT = ("gold_reuse", "derived")  # 这两类有 WT 全长 -> 产 9mer 子肽×HLA
    sub_rows = []
    rerun_no_wt = []  # 待补跑里无 WT 的 (ambiguous/indel_NA)
    for r in rerun.itertuples(index=False):
        pep_id = str(r.Peptide_ID)
        st = status_map.get(pep_id, "indel_NA")
        if st not in HAS_WT:
            rerun_no_wt.append((pep_id, st))
            continue
        wt_full = wt_full_map.get(pep_id, "")
        if not wt_full:
            rerun_no_wt.append((pep_id, st))
            continue
        raw_alleles = "" if pd.isna(r.hla_alleles_to_run) else str(r.hla_alleles_to_run)
        alleles = [a.strip() for a in raw_alleles.split(";") if a.strip()]  # 同患者待补等位(源自 patient_hla)
        if not alleles:
            continue
        mt_pep = "" if pd.isna(r.Vaccine_Peptide) else str(r.Vaccine_Peptide).strip()
        for w in windows:
            for subpep, pos in slide(wt_full, w):  # 复用 MT 同口径滑窗
                for allele in alleles:
                    sub_rows.append({
                        "mut_key": r.mut_key,
                        "Patient_ID": int(r.Patient_ID),
                        "Peptide_ID": pep_id,
                        "Vaccine_Peptide": mt_pep,   # MT 全长, 溯源
                        "WT_FullPeptide": wt_full,
                        "subpep_seq": subpep,        # WT 子肽
                        "subpep_pos": pos,
                        "window_size": w,
                        "hla_allele_std": allele,
                        "side": "WT",
                    })
    sub_df = pd.DataFrame(sub_rows)
    print(f"\n[info] 待补跑肽 WT 子肽展开(gold_reuse+derived): "
          f"{sub_df['Peptide_ID'].nunique() if not sub_df.empty else 0} 肽 -> {len(sub_df)} 行")
    if rerun_no_wt:
        print(f"[info] 待补跑里 {len(rerun_no_wt)} 肽无 WT(ambiguous/indel_NA), 不产 WT 子肽: {rerun_no_wt}")

    # ── 写出 ─────────────────────────────────────────────────────────────
    FROZEN_DIR.mkdir(parents=True, exist_ok=True)
    full_df.to_csv(OUT_WT_FULL, index=False, encoding="utf-8")
    print(f"\n[saved] {OUT_WT_FULL}  shape={full_df.shape}")
    na_df.to_csv(OUT_NA_LIST, index=False, encoding="utf-8")
    print(f"[saved] {OUT_NA_LIST}  shape={na_df.shape}")
    if not sub_df.empty:
        sub_df.to_csv(OUT_WT_SUBPEP, index=False, encoding="utf-8")
        print(f"[saved] {OUT_WT_SUBPEP}  shape={sub_df.shape}")
    else:
        print(f"[WARN] WT 子肽展开为空, 未写 {OUT_WT_SUBPEP}")

    # ── 最终校验门汇总 ───────────────────────────────────────────────────
    sum_snv = n_gold + n_derived + n_amb
    print("\n========== 校验门汇总 ==========")
    print(f"[WT-g1] SNV 数 = {n_snv} (期望 102) {'PASS' if n_snv == 102 else 'CHECK'}; "
          f"gold_reuse+derived+ambiguous = {sum_snv} "
          f"{'PASS' if sum_snv == n_snv else 'CHECK'}")
    print(f"[WT-g4] gold_reuse={n_gold}  derived={n_derived}  "
          f"ambiguous={n_amb}{' PASS(=0)' if n_amb == 0 else ' CHECK(>0, 见上清单)'}  "
          f"indel_NA={len(na_df)}")
    print(f"[WT-g3] 无 WT 数 = {len(na_df)} (期望 28) {'PASS' if len(na_df) == 28 else 'CHECK'}")
    print("[DONE] reconstruct_wt_official 完成 (回推法 match 率见上 [WT-g2])")


if __name__ == "__main__":
    main()
